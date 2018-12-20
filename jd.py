from bs4 import BeautifulSoup
from urllib import request
from urllib.parse import quote
#import openpyxl#操作excel
from openpyxl import Workbook
import time

val=input("请输入关键字：")
kr=quote(str(val))

# url="https://search.jd.com/Search?keyword=%E7%BD%91%E7%BA%BF&enc=utf-8&wq=%E7%BD%91%E7%BA%BF" 
url="https://search.jd.com/Search?keyword="+kr+"&enc=utf-8&wq="+kr 
response=request.urlopen(url)
html=response.read();
html=html.decode("utf-8")
bs=BeautifulSoup(html,"html.parser")
data_list=bs.find_all("li",class_="gl-item")
index=1
wb=Workbook()#创建一个工作簿及一张工作表sheet
#ws=wb.create_sheet(0)#新建sheet到第一的位置
ws=wb.active #获取第一个sheet
ws.title="京东_"+val
ws.cell(row=index,column=1).value="索引"
ws.cell(row=index,column=2).value="标题"
ws.cell(row=index,column=3).value="单价"
ws.cell(row=index,column=4).value="评价"
ws.cell(row=index,column=5).value="类型"
ws.cell(row=index,column=6).value="地址"

for data in data_list:
	index=index+1
	title=data.find("div",class_="p-name p-name-type-2").text.strip()
	price=data.find("div",class_="p-price").text.strip()
	pingjia=data.find("div",class_="p-commit").text.strip()
	leixing=data.find("div",class_="p-icons").text.strip()
	lianjie=data.find("a").get('href')
	ws.cell(row=index,column=1).value=index-1
	ws.cell(row=index,column=2).value=title
	ws.cell(row=index,column=3).value=price
	ws.cell(row=index,column=4).value=pingjia
	ws.cell(row=index,column=5).value=leixing
	ws.cell(row=index,column=6).value=lianjie
	print("\r\n标题："+title+"\r\n价格："+price+"\r\n评价："+pingjia+"\r\n类型："+leixing+"\r\n当前索引："+str(index))
#保存excel
#获得当前时间时间戳 
now = int(time.time()) 
#转换为其他日期格式,如:"%Y-%m-%d %H:%M:%S" 
timeStruct = time.localtime(now) 
strTime = time.strftime("%Y%m%d%H%M%S", timeStruct) 
wb.save("D:\\python\\"+val+strTime+".xlsx")


